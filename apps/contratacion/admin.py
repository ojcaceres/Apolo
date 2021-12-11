import openpyxl
from django.forms import forms
from django.shortcuts import redirect, render
from django.urls import path
from import_export.admin import ExportMixin
from import_export import resources, fields
from django.contrib import admin
from django.contrib import messages
from import_export.widgets import ForeignKeyWidget
from apps.estructura.models import Estado
from apps.autenticacion.models import USUARIO
from .models import  Proceso, Importar, UsuariosProcesos, ModificacionesProcesos


class ProcesoResource(resources.ModelResource):
    username_usuario_asignado = fields.Field(column_name='username_usuario_asignado', attribute='usuario',
                                             widget=ForeignKeyWidget(USUARIO, 'username'))
    nombre_usuario_asignado = fields.Field(column_name='nombre_usuario_asignado', attribute='usuario',
                                           widget=ForeignKeyWidget(USUARIO, 'first_name'))
    username_usuario2_asignado = fields.Field(column_name='Estado', attribute='estado',
                                             widget=ForeignKeyWidget(Estado, 'nombre'))


    class Meta:
        model = Proceso
        import_id_fields = ['usuario', 'estado' ]


class ProcesoUsuarioResource(resources.ModelResource):
    class Meta:
        model = UsuariosProcesos
        import_id_fields = ['usuario', ]


@admin.register(UsuariosProcesos)
class AsignacionUsuariosProcesosAdmin(ExportMixin, admin.ModelAdmin, ):
    resource_class = ProcesoUsuarioResource
    search_fields = ['proceso__numero_de_proceso', 'usuario__username']
    list_display = ['proceso', 'nombre_usuario', 'estado', 'date']
    readonly_fields = ['date']
    raw_id_fields = ['proceso', 'usuario']
    list_filter = ['estado']
    ordering = ('-date',)

class ModificacionProcesoResource(resources.ModelResource):
    class Meta:
        model = ModificacionesProcesos
        import_id_fields = ['proceso', ]

@admin.register(ModificacionesProcesos)
class AsignacionModificacionesProcesosAdmin(ExportMixin, admin.ModelAdmin, ):
    resource_class = ModificacionProcesoResource
    search_fields = ['proceso__numero_de_proceso']
    list_display = ['proceso', 'modificacion']


class UsuariosAsignadosPorProcesoInline(admin.StackedInline):
    model = UsuariosProcesos
    max_num = 1
    readonly_fields = ['date', 'usuario', 'modulo','estado']


@admin.register(Proceso)
class ProcesoAdmin(ExportMixin, admin.ModelAdmin):
    inlines = [UsuariosAsignadosPorProcesoInline]
    resource_class = ProcesoResource
    search_fields = ['cedula_contratista', 'numero_de_proceso', 'nombre_contratista',
                     'numero_contrato']
    raw_id_fields = ['dependencia_contratista', 'proyecto', 'proyecto', 'tipologia_especifica']
    list_filter = ("estado", 'usuario', 'dependencia_contratista', 'proyecto', 'tipologia_especifica')
    list_display = [
        "numero_de_proceso",
        "estado",
        "nombre_contratista",
        "numero_crp",
        "numero_cdp",
        "ultima_actualizacion",
        "usuario",
    ]


class CsvImportForm(forms.Form):
    csv_file = forms.FileField()


def decode_utf8(input_iterator):
    for l in input_iterator:
        yield l.decode('utf-8')

@admin.register(Importar)
class ImportarAdmin(admin.ModelAdmin):
    list_display = ['nombre', 'tipo', 'date']
    change_list_template = "entities/heroes_changelist.html"
    list_filter = ['tipo', 'date']
    search_fields = ['nombre']
    readonly_fields = ['logs', 'date', 'nombre', 'tipo']

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-csv/', self.import_csv),
        ]
        return my_urls + urls

    def import_csv(self, request):
        if request.method == "POST":
            if request.POST.get('tipo') == 'usuario-procesos':
                tipo = 'usuario-procesos'
                excel_file = request.FILES["csv_file"]
                wb = openpyxl.load_workbook(excel_file)
                worksheet = wb.worksheets[0]

                excel_data = list()
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)
                logs = ""
                for i, row in enumerate(excel_data):
                    try:
                        proceso = Proceso.objects.get(numero_de_proceso=row[1])
                        usuario = USUARIO.objects.get(username=row[2])
                        for estado in usuario.estado.all():
                            UsuariosProcesos.objects.create(
                                proceso=proceso,
                                usuario=usuario,
                                estado=estado
                            )
                            logs = logs + f'{usuario.username} asociado a {proceso.numero_de_proceso} perfectamente\n'
                    except Exception as e:
                        logs = logs + f'proceso:{row[1]}, usuario: {row[2]} - Linea: {i} {str(e)}\n'
            elif request.POST.get('tipo') == 'Radicacion':
                tipo = 'Radicacion'

                excel_file = request.FILES["csv_file"]
                wb = openpyxl.load_workbook(excel_file)
                worksheet = wb.worksheets[0]

                excel_data = list()
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)

                logs = ""
                estado_revision = Estado.objects.get(nombre="REVISION")

                for i, row in enumerate(excel_data):
                    try:
                        usuario = USUARIO.objects.get(username=row[9])
                        Proceso.objects.create(
                            numero_de_proceso=row[0].strip(),
                            nombre_contratista=row[1].strip(),
                            cedula_contratista=row[2].strip(),
                            numero_cdp=row[3].strip(),
                            valor_cdp=row[4].strip(),
                            valor_contrato=row[5].strip(),
                            proyecto_id=int(row[6].strip()),
                            objeto=row[7].strip(),
                            usuario=usuario,
                            estado=estado_revision
                        )
                        linea_log = f'Proceso creado'

                        logs = logs + f'Linea: {i} - {linea_log}\n'
                    except Exception as e:
                        logs = logs + f'Linea: {i} {str(e)}\n'

            Importar.objects.create(nombre=request.FILES["csv_file"].name, tipo=tipo, logs=logs)

            self.message_user(request=request, message='El archivo se ha subido exitosamente', level=messages.INFO)
            return redirect("..")
        form = CsvImportForm()
        context = dict(
            self.admin_site.each_context(request),
            form=form,
        )
        return render(
            request, "admin/csv_form.html", context)
